from fastapi import APIRouter, Depends, HTTPException, Query, status, UploadFile, File
from sqlalchemy.orm import Session
from typing import Callable, List, Optional
from datetime import date, datetime
import io

from app.database import get_db
from app.schemas.student import (
    StudentCreate,
    StudentUpdate,
    StudentResponse,
    StudentProgressCreate,
    StudentProgressResponse,
)
from app.models.math_class import MathClass
from app.models.student import Student
from app.models.worksheet import Worksheet
from app.utils.dependencies import get_current_teacher
from app.models.user import User
from app.services import student_service


router = APIRouter()

EXPECTED_EXCEL_HEADERS = [
    "Họ và tên",
    "Ngày sinh",
    "Họ tên bố/mẹ",
    "SĐT bố/mẹ",
]


def _clean_text(value: object) -> Optional[str]:
    if value is None:
        return None
    cleaned = str(value).strip()
    return cleaned or None


def _parse_dob(
    raw_value: object,
    row_index: int,
    excel_date_parser: Optional[Callable[[float], datetime | date]] = None,
) -> Optional[date]:
    """Parse DOB from Excel values in multiple common formats."""
    if raw_value is None:
        return None

    if isinstance(raw_value, datetime):
        return raw_value.date()
    if isinstance(raw_value, date):
        return raw_value

    if isinstance(raw_value, (int, float)):
        if excel_date_parser is None:
            raise HTTPException(
                status_code=500,
                detail="Thiếu thư viện xử lý Excel (openpyxl). Vui lòng cài đặt lại dependencies backend.",
            )
        try:
            parsed = excel_date_parser(raw_value)
            return parsed.date() if isinstance(parsed, datetime) else parsed
        except Exception as exc:
            raise HTTPException(
                status_code=400,
                detail=f"Ngày sinh không hợp lệ ở dòng {row_index}"
            ) from exc

    value = str(raw_value).strip()
    if not value:
        return None

    formats = ["%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d/%m/%y", "%d.%m.%Y"]
    for date_format in formats:
        try:
            return datetime.strptime(value, date_format).date()
        except ValueError:
            continue

    raise HTTPException(
        status_code=400,
        detail=f"Không đọc được ngày sinh ở dòng {row_index}: '{value}'"
    )


def verify_class_ownership(db: Session, class_id: int, teacher_id: int) -> MathClass:
    """Verify that the teacher owns the class."""
    db_class = db.query(MathClass).filter(MathClass.id == class_id).first()
    if not db_class:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Không tìm thấy lớp học"
        )
    if int(db_class.teacher_id) != teacher_id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Bạn không có quyền truy cập lớp học này"
        )
    return db_class


@router.get("/classes/{class_id}/students", response_model=List[StudentResponse])
async def list_students(
    class_id: int,
    tier: Optional[str] = None,
    skip: int = Query(0, ge=0),
    limit: int = Query(20, ge=1, le=100),
    current_user: User = Depends(get_current_teacher),
    db: Session = Depends(get_db)
):
    """
    Lấy danh sách học sinh trong lớp.
    
    - **tier**: Lọc theo nhóm (foundation, standard, extension, advanced)
    """
    verify_class_ownership(db, class_id, int(current_user.id))
    return student_service.get_students_by_class(db, class_id, tier, skip=skip, limit=limit)


@router.post("/classes/{class_id}/students", response_model=StudentResponse, status_code=status.HTTP_201_CREATED)
async def create_student(
    class_id: int,
    student_data: StudentCreate,
    current_user: User = Depends(get_current_teacher),
    db: Session = Depends(get_db)
):
    """
    Thêm học sinh vào lớp.
    
    - **full_name**: Họ tên học sinh
    - **dob**: Ngày tháng năm sinh (YYYY-MM-DD)
    - **parent_name**: Họ tên bố hoặc mẹ
    - **parent_phone**: SĐT bố hoặc mẹ
    - **tier**: Nhóm năng lực (foundation/standard/extension/advanced)
    """
    verify_class_ownership(db, class_id, int(current_user.id))
    
    tier_val = student_data.tier.value if student_data.tier else "standard"
    return student_service.create_student(
        db,
        student_data.full_name,
        class_id,
        tier_val,
        student_data.dob,
        student_data.parent_name,
        student_data.parent_phone,
    )


@router.post("/classes/{class_id}/students/upload", response_model=List[StudentResponse], status_code=status.HTTP_201_CREATED)
async def upload_students_excel(
    class_id: int,
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_teacher),
    db: Session = Depends(get_db)
):
    """
    Thêm học sinh từ file Excel theo đúng mẫu cột tiếng Việt.
    """
    verify_class_ownership(db, class_id, int(current_user.id))
    
    if not file.filename or not file.filename.lower().endswith('.xlsx'):
        raise HTTPException(status_code=400, detail="Chỉ hỗ trợ định dạng .xlsx")
        
    try:
        try:
            import openpyxl
            from openpyxl.utils.datetime import from_excel
        except ModuleNotFoundError as exc:
            raise HTTPException(
                status_code=500,
                detail="Máy chủ thiếu thư viện openpyxl nên chưa thể import Excel. Vui lòng chạy: pip install -r backend/requirements.txt",
            ) from exc

        contents = await file.read()
        workbook = openpyxl.load_workbook(io.BytesIO(contents))
        sheet = workbook.active
        if sheet is None:
            raise HTTPException(status_code=400, detail="File Excel khong co sheet hop le")

        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        actual_headers = [str(cell).strip() if cell is not None else "" for cell in (header_row or [])]
        if actual_headers[:4] != EXPECTED_EXCEL_HEADERS:
            expected = ", ".join(EXPECTED_EXCEL_HEADERS)
            raise HTTPException(
                status_code=400,
                detail=f"File Excel sai mẫu. Cần đúng tiêu đề: {expected}"
            )
        
        students_to_create: List[Student] = []
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            cells = list(row or [])
            while len(cells) < 4:
                cells.append(None)

            name = _clean_text(cells[0])
            if not name:
                continue

            dob_value = _parse_dob(cells[1], row_idx, from_excel)
            parent_name = _clean_text(cells[2])
            parent_phone = _clean_text(cells[3])

            students_to_create.append(
                Student(
                    full_name=name,
                    class_id=class_id,
                    tier="standard",
                    dob=dob_value,
                    parent_name=parent_name,
                    parent_phone=parent_phone,
                )
            )

        if not students_to_create:
            raise HTTPException(status_code=400, detail="Không có học sinh hợp lệ để import")

        return student_service.bulk_create_students(db, students_to_create)
    except Exception as e:
        if isinstance(e, HTTPException):
            raise e
        raise HTTPException(status_code=400, detail=f"Lỗi đọc file Excel: {str(e)}")


@router.get("/students/{student_id}", response_model=StudentResponse)
async def get_student(
    student_id: int,
    current_user: User = Depends(get_current_teacher),
    db: Session = Depends(get_db)
):
    """
    Lấy thông tin học sinh.
    """
    student = student_service.get_student_by_id(db, student_id)
    if not student:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Không tìm thấy học sinh"
        )
    
    verify_class_ownership(db, int(student.class_id), int(current_user.id))
    return student


@router.put("/students/{student_id}", response_model=StudentResponse)
async def update_student(
    student_id: int,
    student_data: StudentUpdate,
    current_user: User = Depends(get_current_teacher),
    db: Session = Depends(get_db)
):
    """
    Cập nhật thông tin học sinh.
    """
    student = student_service.get_student_by_id(db, student_id)
    if not student:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Không tìm thấy học sinh"
        )
    
    verify_class_ownership(db, int(student.class_id), int(current_user.id))
    
    tier_val = student_data.tier.value if student_data.tier else None
    return student_service.update_student(
        db,
        student,
        student_data.full_name,
        tier_val,
        student_data.dob,
        student_data.parent_name,
        student_data.parent_phone,
    )


@router.delete("/students/{student_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_student(
    student_id: int,
    current_user: User = Depends(get_current_teacher),
    db: Session = Depends(get_db)
):
    """
    Xóa học sinh khỏi lớp.
    """
    student = student_service.get_student_by_id(db, student_id)
    if not student:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Không tìm thấy học sinh"
        )
    
    verify_class_ownership(db, int(student.class_id), int(current_user.id))
    student_service.delete_student(db, student)
    return None


@router.post("/students/{student_id}/progress", response_model=StudentProgressResponse)
async def save_student_progress(
    student_id: int,
    payload: StudentProgressCreate,
    current_user: User = Depends(get_current_teacher),
    db: Session = Depends(get_db),
):
    """Lưu tiến độ làm bài của học sinh sau khi giáo viên duyệt kết quả AI."""
    student = student_service.get_student_by_id(db, student_id)
    if not student:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Không tìm thấy học sinh"
        )

    verify_class_ownership(db, int(student.class_id), int(current_user.id))

    worksheet = db.query(Worksheet).filter(Worksheet.id == payload.worksheet_id).first()
    if not worksheet:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Không tìm thấy bài tập"
        )
    if int(worksheet.class_id) != int(student.class_id):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Bài tập không thuộc lớp của học sinh"
        )

    progress_details = dict(payload.details or {})
    progress_details["teacher_approval_status"] = "approved"
    if payload.score is not None:
        progress_details["score"] = payload.score

    return student_service.save_student_progress(
        db,
        student_id=student.id,
        worksheet_id=payload.worksheet_id,
        correct_count=payload.correct_count,
        total_count=payload.total_count,
        details=progress_details,
    )
